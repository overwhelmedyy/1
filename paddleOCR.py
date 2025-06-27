from flask import Blueprint, jsonify, Response
import json
import time
from paddleocr import PaddleOCR
from app.parsers.paddleParser import parse_text
from app.utils.operations import add_invoice_to_db, check_if_invoice, process_paddleocr_text, add_mu_invoice_to_db, add_mu_packing_list_to_db, get_packing_list_with_supplier_invno
from app.utils.utils import load_image, get_files_from_request

import grpc
import grpc_client.service_pb2 as service__pb2
import grpc_client.service_pb2_grpc as service__pb2_grpc

import pandas as pd
from openpyxl import load_workbook
import psycopg2
from psycopg2 import sql


paddleocr_bp = Blueprint('paddleocr', __name__)




from flask import request
import os
SHARED_FOLDER = r'C:\Users\SCSSH190\Desktop\all_the_projects\251\backend\shared-data\input'
#SHARED_FOLDER = '/app/shared-data/input'
INVOICE_FOLDER = 'invoices'
PACKING_LIST_FOLDER = 'packing_list'

@paddleocr_bp.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return "没有文件", 400
    file = request.files['file']
    if file.filename == '':
        return "没有选择文件", 400

    # 保存文件到共享文件夹
    file_path = os.path.join(SHARED_FOLDER, INVOICE_FOLDER)
    os.makedirs(file_path,exist_ok=True)
    file_path = os.path.join(file_path,file.filename)
    file.save(file_path)

    return jsonify({"ret":0})

@paddleocr_bp.route('/upload-packing-list', methods=['POST'])
def upload_packing_list():
    if 'file' not in request.files:
        return "没有文件", 400
    file = request.files['file']
    if file.filename == '':
        return "没有选择文件", 400

    # 保存文件到共享文件夹
    # file_path = os.path.join(SHARED_FOLDER, PACKING_LIST_FOLDER, file.filename)
    # file.save(file_path)
    # 保存文件到共享文件夹
    file_path = os.path.join(SHARED_FOLDER, PACKING_LIST_FOLDER)
    os.makedirs(file_path,exist_ok=True)
    file_path = os.path.join(file_path,file.filename)
    file.save(file_path)

    return jsonify({"ret":0})



@paddleocr_bp.route('/paddleOCR2', methods=['POST'])
def process_paddleocr2():

    temp_folder_path = r'C:\Users\SCSSH190\Desktop\all_the_projects\251\backend\shared-data'

    for file_name in os.listdir(temp_folder_path):

        file_path = os.path.join(temp_folder_path, file_name)


        if not os.path.isfile(file_path):
            print(f"跳过非文件：{file_name}")
            continue

        if file_name.startswith("I"):
            IorP = 0
        elif file_name.startswith("P"):
            IorP = 1
        else:
            print(f"跳过未知类型文件：{file_name}")
            continue

        try:
            # muVision OCR
            with grpc.insecure_channel('127.0.0.1:50051') as channel:
                # 测试 OCR 服务
                stub = service__pb2_grpc.OCRServiceStub(channel)
                req = service__pb2.OCRRequest(file_path=file_path, file_type=IorP)
                res = stub.InferCN(req)
                actual_invoice_rslt_path = os.path.join('shared-data', 'output', os.path.basename(res.result_path))
        except grpc.RpcError as e:
            print(f"GRPC 错误: {e}")
            return Response(json.dumps({"status": "error", "message": str(e)}), mimetype='application/json')

        if IorP == 0:
            most_frequent_invoice_number, is_foreign = I_add_to_db(actual_invoice_rslt_path, res)
            if is_foreign and most_frequent_invoice_number:
                PL_merge2_I(res.supplier, actual_invoice_rslt_path, most_frequent_invoice_number)

        elif IorP == 1:
             PL_add_to_db(actual_invoice_rslt_path, res)

        response = {
            'status': 'success',
            'result_path': res.result_path,
            'supplier': res.supplier
        }
    response_json = json.dumps(response, ensure_ascii=False)
    return Response(response_json, mimetype='application/json')



def PL_add_to_db(actual_packing_list_rslt_path, res):

    df = pd.read_excel(actual_packing_list_rslt_path)
    for index, row in df.iterrows():
        print(row)
        invoice_number = row['发票号']
        number = row['件数']
        gross = row['毛重']
        packing = row['包装']

        add_mu_packing_list_to_db(res.supplier, invoice_number, number, gross, packing)

def I_add_to_db(actual_invoice_rslt_path,res):

    if os.path.isfile(actual_invoice_rslt_path):
        df = pd.read_excel(actual_invoice_rslt_path)

        if df.empty:
            print("识别结果excel为空")
            return Response(json.dumps({"status": "error", "message": "识别结果为空"}), mimetype='application/json')

        if '发票号' not in df.columns:
            print("发票号列不存在")
            return Response(json.dumps({"status": "error", "message": "发票号列不存在"}),
                            mimetype='application/json')

        if df['发票号'].isnull().all():
            print("发票号列全为空")
            return Response(json.dumps({"status": "error", "message": "发票号列全为空"}),
                            mimetype='application/json')

        is_foreign_invoice = False
        most_frequent_invoice_number = str(df['发票号'].value_counts().idxmax())

        for index, row in df.iterrows():
            print(row)
            invoice_number = most_frequent_invoice_number if most_frequent_invoice_number else row['发票号']
            order_number = row['订单号']
            item_name = row['村田品番']
            unit_price = row['单价']
            quantity = row['数量']
            total_price = row['总金额']
            currency = row['币别']
            origin = ""
            if currency != "CNY":
                origin = row['原产国']
                is_foreign_invoice = True

            pdf_file, image_file = get_files_from_request()
            add_mu_invoice_to_db(pdf_file, image_file, res.supplier, invoice_number, order_number, item_name,
                                 unit_price, quantity, total_price, currency, origin)

            return most_frequent_invoice_number, is_foreign_invoice

def PL_merge2_I(supplier, actual_invoice_rslt_path, most_frequent_invoice_number):

    invoice_rslt_excel = load_workbook(actual_invoice_rslt_path)
    worksheet = invoice_rslt_excel['Sheet']
    worksheet['I1'] = '件数'
    worksheet['J1'] = '毛重'
    worksheet['K1'] = '包装'

    packing_lists = get_packing_list_with_supplier_invno(supplier, most_frequent_invoice_number)
    if packing_lists:
        packing_list = packing_lists[0]
        # 修正Invoice识别结果发票号
        for i, x in enumerate(next(worksheet.iter_cols(min_row=2, max_col=0, values_only=True))):
            if type(x) is not str or str(x) != str(most_frequent_invoice_number):
                worksheet['A' + f'{i + 2}'] = str(most_frequent_invoice_number)
        # 写入packing list内容
        worksheet['I2'] = packing_list['number']
        worksheet['J2'] = packing_list['gross']
        worksheet['K2'] = packing_list['packing']



@paddleocr_bp.route('/paddleOCR', methods=['POST'])
def process_paddleocr():
    try:
        print(request)

        # muVision OCR
        with grpc.insecure_channel('127.0.0.1:50051') as channel:
            # 测试 OCR 服务
            stub = service__pb2_grpc.OCRServiceStub(channel)

            # 准备请求数据
            print(f"Request: {request.json}")
            req = service__pb2.OCRRequest(image_paths=request.json["file_name"], file_type=request.json["packing_list_name"])
            print(req)
            res = stub.InferCN(req)
            print("识别结果路径:", res.result_path)
            print("Packing list识别结果路径:", res.packing_list_result_path)
            print("供应商为:", res.supplier)

            supplier = res.supplier

            # 如果有Packing list，存储Packing List识别结果至数据库
            actual_packing_list_rslt_path = os.path.join('shared-data', 'output', os.path.basename(res.packing_list_result_path))
            if os.path.isfile(actual_packing_list_rslt_path):
                df = pd.read_excel(actual_packing_list_rslt_path)

                for index, row in df.iterrows():
                    print(row)
                    invoice_number = row['发票号']
                    number = row['件数']
                    gross = row['毛重']
                    packing = row['包装']

                    add_mu_packing_list_to_db(supplier, invoice_number, number, gross, packing)

            # 存储Invoice识别结果至数据库
            actual_invoice_rslt_path = os.path.join('shared-data', 'output', os.path.basename(res.result_path))
            if os.path.isfile(actual_invoice_rslt_path):
                df = pd.read_excel(actual_invoice_rslt_path)

                if df.empty:
                    print("识别结果excel为空")
                    return Response(json.dumps({"status": "error", "message": "识别结果为空"}), mimetype='application/json')

                if '发票号' not in df.columns:
                    print("发票号列不存在")
                    return Response(json.dumps({"status": "error", "message": "发票号列不存在"}), mimetype='application/json')

                if df['发票号'].isnull().all():
                    print("发票号列全为空")
                    return Response(json.dumps({"status": "error", "message": "发票号列全为空"}), mimetype='application/json')

                is_foreign_invoice = False
                most_frequent_invoice_number = str(df['发票号'].value_counts().idxmax())

                for index, row in df.iterrows():
                    print(row)
                    invoice_number = most_frequent_invoice_number if most_frequent_invoice_number else row['发票号']
                    order_number = row['订单号']
                    item_name = row['村田品番']
                    unit_price = row['单价']
                    quantity = row['数量']
                    total_price = row['总金额']
                    currency = row['币别']
                    origin = ""
                    if currency != "CNY":
                        origin = row['原产国']
                        is_foreign_invoice = True

                    pdf_file, image_file = get_files_from_request()
                    add_mu_invoice_to_db(pdf_file, image_file, supplier, invoice_number, order_number, item_name, unit_price, quantity, total_price, currency, origin)

                invoice_rslt_excel = load_workbook(actual_invoice_rslt_path)
                # 如果是外币发票，将Packing List数据合并至Invoice识别结果excel，同时修正发票号
                if is_foreign_invoice and most_frequent_invoice_number:
                    # 添加'件数'，'毛重'，'包装'表头
                    worksheet = invoice_rslt_excel['Sheet']
                    worksheet['I1'] = '件数'
                    worksheet['J1'] = '毛重'
                    worksheet['K1'] = '包装'

                    packing_lists = get_packing_list_with_supplier_invno(supplier, most_frequent_invoice_number)
                    if packing_lists:
                        packing_list = packing_lists[0]
                        # 修正Invoice识别结果发票号
                        for i, x in enumerate(next(worksheet.iter_cols(min_row=2, max_col=0, values_only=True))):
                            if type(x) is not str or str(x) != str(most_frequent_invoice_number):
                                worksheet['A' + f'{i + 2}'] = str(most_frequent_invoice_number)
                        # 写入packing list内容
                        worksheet['I2'] = packing_list['number']
                        worksheet['J2'] = packing_list['gross']
                        worksheet['K2'] = packing_list['packing']

                invoice_rslt_excel.save(actual_invoice_rslt_path)

            # 返回成功响应
            response = {
                'status': 'success',
                'result_path': res.result_path,
                'supplier': res.supplier
            }
            response_json = json.dumps(response, ensure_ascii=False)
            return Response(response_json, mimetype='application/json')

    except grpc.RpcError as e:
        print(f"GRPC 错误: {e}")
        return Response(json.dumps({"status": "error", "message": str(e)}), mimetype='application/json')
    except Exception as e:
        print(f"发生错误: {e}")
        return Response(json.dumps({"status": "error", "message": str(e)}), mimetype='application/json')

        # return jsonify(response, ensure_ascii=False)


    # Original
    # ocr = PaddleOCR(
    #     det_model_dir='paddle_models/en_PP-OCRv3_det_infer',
    #     rec_model_dir='paddle_models/en_PP-OCRv3_rec_infer',
    #     cls_model_dir='paddle_models/ch_ppocr_mobile_v2.0_cls_infer',
    #     use_angle_cls=True,
    #     lang='en'
    # )
    # img = load_image()
    # ocr_method = 'PaddleOCR'

    # start_time_recognition = time.time()
    # result = ocr.ocr(img, cls=True)
    # recognition_time = time.time() - start_time_recognition

    # average_confidence, text = process_paddleocr_text(result)

    # start_time_parsing = time.time()
    # parsed_data = parse_text(text)
    # parsing_time = time.time() - start_time_parsing

    # response = {
    #     'text': text,
    #     'parsed_data': parsed_data,
    #     'time': {
    #         'recognition': recognition_time,
    #         'parsing': parsing_time,
    #     },
    #     'average_confidence': average_confidence * 100
    # }

    # if check_if_invoice(parsed_data):
    #     pdf_file, image_file = get_files_from_request()
    #     invoice_id = add_invoice_to_db(parsed_data, text, pdf_file, image_file,
    #                                    average_confidence * 100, recognition_time, parsing_time, ocr_method)
    #     response['invoice_id'] = invoice_id

    # return jsonify(response)
